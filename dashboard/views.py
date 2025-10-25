# dashboard/views.py
import base64
import os
import shutil
import tempfile
from textwrap import wrap
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required , user_passes_test
from django.db.models import Sum, Count, Avg
from django.http import HttpResponse
from django.db import transaction # Pour gérer les transactions de base de données
from django.forms import ValidationError, modelformset_factory # Pour gérer plusieurs formulaires d'un même modèle
from django.forms import formset_factory
from django.utils import timezone
# Pour la génération de PDF
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from datetime import datetime, timedelta
from datetime import date
from reportlab.lib.units import inch
from django.core.files.base import ContentFile
import io
import os # Pour extraire le nom de base du fichier d'asset
import qrcode
from weasyprint import HTML
from wkhtmltopdf.views import PDFTemplateResponse
from django.template.loader import render_to_string
import pdfkit  
from django_pdfkit import PDFView

from reportlab.lib.pagesizes import A5
from reportlab.lib.utils import ImageReader
from reportlab.lib.colors import HexColor, black
from django.db.models import Avg, Min, Max
import csv
import openpyxl

# Importation de ReportLab
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.utils import ImageReader



from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Table, TableStyle

# Importez tous vos formulaires
from .forms import (
    CertificatFrequentationForm, EcoleSettingsForm, EmploiDuTempsForm, EtudiantForm, DossierInscriptionImageForm, NoteForm, PaiementForm, PresenceForm,
    EnseignantForm, MatiereForm, ClasseForm, ProgrammeMatiereForm, AnneeScolaireForm
)

# Importez tous vos modèles
from .models import (
    EcoleSettings, EmploiDuTemps, Etudiant, AnneeScolaire, Enseignant, Classe, Matiere, ProgrammeMatiere,
    Note, Paiement, Presence, DossierInscriptionImage, CertificatFrequentation , Profile
)

from django.db.models import Q  # Pour les recherches complexes

# ------------------------------------------------------------------
# UTILITAIRE : récupérer l'école de l'utilisateur connecté
# ------------------------------------------------------------------



@login_required
def initier_paiement(request):
    """
    Vue simulant un paiement pour l'école associée à l'utilisateur.
    """
    ecole = getattr(request.user.profile, 'ecole', None)

    if not ecole:
        messages.error(request, "Aucune école associée à votre compte.")
        return redirect('dashboard_accueil')

    # Si l'école est déjà active, inutile de payer à nouveau
    if ecole.est_active and not ecole.periode_essai_expiree():
        messages.info(request, "Votre abonnement est déjà actif.")
        return redirect('dashboard_accueil')

    if request.method == "POST":
        # Simuler la réussite du paiement (à remplacer par l’API réelle)
        ecole.est_active = True
        ecole.date_fin_essai = timezone.now() + timedelta(days=30)  # 1 mois d’abonnement
        ecole.save()

        messages.success(request, "Paiement effectué avec succès ! Votre abonnement a été renouvelé.")
        return redirect('dashboard_accueil')

    context = {
        'ecole': ecole,
        'montant': 25000,  # Exemple : 5000 FCFA / mois
    }
    return render(request, 'dashboard/paiements/initier_paiement.html', context)




def get_user_ecole(request):
    """Retourne l'école associée au profil utilisateur connecté."""
    if hasattr(request.user, 'profile') and request.user.profile.ecole:
        return request.user.profile.ecole
    return None


# ------------------------------------------------------------------
# RECHERCHE D'ÉTUDIANTS PAR ÉCOLE
# ------------------------------------------------------------------
@login_required
def recherche_etudiants(request):
    query = request.GET.get('q', '').strip()
    ecole = get_user_ecole(request)
    resultats = []

    if not ecole:
        messages.warning(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    if query:
        resultats = (
            Etudiant.objects.filter(ecole=ecole)
            .filter(
                Q(nom__icontains=query)
                | Q(prenom__icontains=query)
                | Q(numero_matricule__icontains=query)
                | Q(classe__nom_classe__icontains=query)
            )
            .select_related('classe', 'annee_scolaire_inscription')
        )

    context = {
        'query': query,
        'resultats': resultats,
    }
    return render(request, 'dashboard/etudiants/recherche_etudiants.html', context)


# ------------------------------------------------------------------
# TABLEAU DE BORD ACCUEIL (MULTI-ÉCOLE)
# ------------------------------------------------------------------
@login_required
def dashboard_accueil(request):
    ecole = get_user_ecole(request)

    # 🔒 Vérifie que l'utilisateur est lié à une école
    if not ecole:
        messages.warning(request, "Vous n'êtes associé à aucune école.")
        return redirect('choisir_ecole')  # ou une vue pour sélectionner/créer une école

    # 🏫 Année scolaire active pour CETTE école uniquement
    annee_active = AnneeScolaire.objects.filter(ecole=ecole, active=True).first()

    if not annee_active:
        messages.warning(request, f"Aucune année scolaire active n'est définie pour l'école {ecole.nom_etablissement}.")
        return redirect('liste_annees_scolaires')

    # 💰 Élèves avec paiements impayés ou partiels
    eleves_non_payes_compte = (
        Paiement.objects.filter(
            annee_scolaire=annee_active,
            ecole=ecole,
            statut__in=['Impayé', 'Partiel']
        )
        .values('etudiant')
        .annotate(total_du_par_etudiant=Sum('montant_du'))
        .count()
    )

    # 🚫 Absents aujourd'hui
    absents_aujourd_hui = Presence.objects.filter(
        date=timezone.now().date(),
        annee_scolaire=annee_active,
        ecole=ecole,
        statut='Absent'
    ).count()

    # ✅ Total payé
    total_paye = (
        Paiement.objects.filter(
            annee_scolaire=annee_active,
            ecole=ecole,
            statut__in=['Payé', 'Partiel']
        ).aggregate(total_paid=Sum('montant'))['total_paid']
        or 0
    )

    # ❌ Total impayé
    total_impaye = (
        Paiement.objects.filter(
            annee_scolaire=annee_active,
            ecole=ecole,
            statut__in=['Impayé', 'Partiel']
        ).aggregate(total_due=Sum('montant_du'))['total_due']
        or 0
    )

    # 🧑‍🎓 3 élèves récemment inscrits
    etudiants_recents = (
        Etudiant.objects.filter(
            ecole=ecole,
            annee_scolaire_inscription=annee_active
        )
        .order_by('-date_inscription')[:3]
    )

    # 🏫 Compter classes & enseignants
    classes_actives = Classe.objects.filter(annee_scolaire=annee_active, ecole=ecole).count()
    enseignants_actifs = ecole.enseignants.count()
 # relation inverse depuis modèle Enseignant

    context = {
        'annee_active': annee_active,
        'ecole_active': ecole,
        'eleves_non_payes_compte': eleves_non_payes_compte,
        'absents_aujourd_hui': absents_aujourd_hui,
        'total_paye': total_paye,
        'total_impaye': total_impaye,
        'etudiants_recents': etudiants_recents,
        'classes_actives': classes_actives,
        'enseignants_actifs': enseignants_actifs,
    }

    return render(request, 'dashboard/dashboard_accueil.html', context)

# ------------------------------------------------------------------
# LISTE ÉTUDIANTS
# ------------------------------------------------------------------
@login_required
def liste_etudiants(request):
    ecole = get_user_ecole(request)
    etudiants = Etudiant.objects.filter(ecole=ecole).select_related('classe', 'annee_scolaire_inscription')
    toutes_les_classes = Classe.objects.filter(ecole=ecole).order_by('nom_classe')

    selected_classe_id = request.GET.get('classe')
    if selected_classe_id:
        try:
            selected_classe_id = int(selected_classe_id)
            etudiants = etudiants.filter(classe__id=selected_classe_id)
        except (ValueError, TypeError):
            messages.warning(request, "Le filtre de classe sélectionné n'est pas valide.")
            selected_classe_id = None

    etudiants = etudiants.order_by('classe__nom_classe', 'nom', 'prenom')

    context = {
        'etudiants': etudiants,
        'toutes_les_classes': toutes_les_classes,
        'selected_classe_id': selected_classe_id
    }
    return render(request, 'dashboard/etudiants/liste_etudiants.html', context)

# ------------------------------------------------------------------
# CRÉER ÉTUDIANT
# ------------------------------------------------------------------
@login_required
def creer_etudiant(request):
    # Récupérer l'école de l'utilisateur
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # Formset pour les images du dossier d'inscription
    DossierFormSet = modelformset_factory(
        DossierInscriptionImage,
        form=DossierInscriptionImageForm,
        extra=3,
        can_delete=True
    )

    # Récupérer l'année scolaire active pour cette école
    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()

    if request.method == 'POST':
        form = EtudiantForm(
            request.POST,
            request.FILES,
            ecole=ecole,  # ⚡ Passer l'école pour filtrer classes et années
        )
        formset = DossierFormSet(
            request.POST,
            request.FILES,
            queryset=DossierInscriptionImage.objects.none()
        )

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                # Créer l'étudiant
                etudiant = form.save(commit=False)
                etudiant.ecole = ecole
                # Assigner l'année active si non fournie
                if not etudiant.annee_scolaire_inscription:
                    etudiant.annee_scolaire_inscription = annee_active
                etudiant.save()

                # Sauvegarde des images du dossier
                for f in formset:
                    if f.cleaned_data and not f.cleaned_data.get('DELETE'):
                        dossier_image = f.save(commit=False)
                        dossier_image.etudiant = etudiant
                        dossier_image.save()

                messages.success(
                    request,
                    f"L'élève {etudiant.prenom} {etudiant.nom} a été ajouté avec succès."
                )
                return redirect('detail_etudiant', etudiant_id=etudiant.pk)
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = EtudiantForm(
            ecole=ecole,
            initial={'annee_scolaire_inscription': annee_active}
        )
        formset = DossierFormSet(queryset=DossierInscriptionImage.objects.none())

    return render(
        request,
        'dashboard/etudiants/creer_etudiant.html',
        {'form': form, 'formset': formset, 'annee_active': annee_active}
    )


# ------------------------------------------------------------------
# DÉTAIL ÉTUDIANT
# ------------------------------------------------------------------
@login_required
def detail_etudiant(request, etudiant_id):
    # Récupérer l'école de l'utilisateur connecté
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # Étudiant filtré par école
    etudiant = get_object_or_404(
        Etudiant.objects.select_related('classe', 'annee_scolaire_inscription'),
        pk=etudiant_id,
        ecole=ecole
    )

    # Année scolaire active pour cette école
    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()

    # Dossier d'inscription
    dossier_images = etudiant.dossier_images.all()

    # Notes filtrées par école et année active (ou tout l'historique si besoin)
    notes = etudiant.notes.filter(
        ecole=ecole,
        annee_scolaire=annee_active
    ).select_related('matiere', 'annee_scolaire').order_by('-annee_scolaire', 'periode_evaluation')

    # Paiements filtrés par école
    paiements = etudiant.paiements.filter(
        ecole=ecole,
        annee_scolaire=annee_active
    ).select_related('annee_scolaire').order_by('-date_paiement')

    # Présences filtrées par école et année active
    presences = etudiant.presences.filter(
        ecole=ecole,
        annee_scolaire=annee_active
    ).select_related('matiere', 'annee_scolaire').order_by('-date')

    # Calculs financiers
    total_paye = paiements.filter(statut__in=['Payé', 'Partiel']).aggregate(Sum('montant'))['montant__sum'] or 0
    total_du = paiements.aggregate(Sum('montant_du'))['montant_du__sum'] or 0
    solde_restant = total_du - total_paye

    # Statistiques de présence
    stats_presences = presences.values('statut').annotate(total=Count('id'))
    total_jours_presents = next((x['total'] for x in stats_presences if x['statut'] == 'Présent'), 0)
    total_jours_absents = next((x['total'] for x in stats_presences if x['statut'] == 'Absent'), 0)
    total_jours_retard = next((x['total'] for x in stats_presences if x['statut'] == 'Retard'), 0)
    total_jours_excuses = next((x['total'] for x in stats_presences if x['statut'] == 'Excusé'), 0)

    context = {
        'etudiant': etudiant,
        'dossier_images': dossier_images,
        'notes': notes,
        'paiements': paiements,
        'presences': presences,
        'total_paye': total_paye,
        'total_du': total_du,
        'solde_restant': solde_restant,
        'total_jours_presents': total_jours_presents,
        'total_jours_absents': total_jours_absents,
        'total_jours_retard': total_jours_retard,
        'total_jours_excuses': total_jours_excuses,
        'annee_active': annee_active,
    }

    return render(request, 'dashboard/etudiants/detail_etudiant.html', context)




# ------------------------------------------------------------------
# MODIFIER ÉTUDIANT
# ------------------------------------------------------------------
@login_required
def modifier_etudiant(request, etudiant_id):
    """
    Modifier un étudiant et ses dossiers d'inscription.
    Filtrage strict par l'école de l'utilisateur.
    """
    # 🔹 Récupération de l'école
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # 🔹 Récupération de l'étudiant lié à cette école
    etudiant = get_object_or_404(Etudiant, pk=etudiant_id, ecole=ecole)

    # 🔹 Formset pour gérer les images/dossiers
    DossierFormSet = modelformset_factory(
        DossierInscriptionImage,
        form=DossierInscriptionImageForm,
        extra=1,
        can_delete=True
    )

    if request.method == 'POST':
        # ⚡ Important : passer l'école au formulaire pour filtrer les classes et années
        form = EtudiantForm(request.POST, request.FILES, instance=etudiant, ecole=ecole)
        formset = DossierFormSet(
            request.POST,
            request.FILES,
            queryset=DossierInscriptionImage.objects.filter(etudiant=etudiant)
        )

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                etudiant = form.save()
                
                for f in formset:
                    if f.cleaned_data:
                        if f.cleaned_data.get('DELETE') and f.instance.pk:
                            f.instance.delete()
                        else:
                            dossier_image = f.save(commit=False)
                            dossier_image.etudiant = etudiant
                            dossier_image.save()
                
                messages.success(request, f"Les informations de {etudiant.prenom} {etudiant.nom} ont été mises à jour.")
                return redirect('detail_etudiant', etudiant_id=etudiant.pk)
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = EtudiantForm(instance=etudiant, ecole=ecole)
        formset = DossierFormSet(queryset=DossierInscriptionImage.objects.filter(etudiant=etudiant))

    return render(request, 'dashboard/etudiants/modifier_etudiant.html', {
        'form': form,
        'formset': formset,
        'etudiant': etudiant
    })


@login_required
def supprimer_etudiant(request, etudiant_id):
    """
    Supprimer un étudiant (filtré par école de l'utilisateur).
    """
    ecole = get_user_ecole(request)
    etudiant = get_object_or_404(Etudiant, pk=etudiant_id, ecole=ecole)

    if request.method == 'POST':
        nom_complet = f"{etudiant.prenom} {etudiant.nom}"
        etudiant.delete()
        messages.success(request, f"L'élève {nom_complet} a été supprimé avec succès.")
        return redirect('liste_etudiants')

    return render(request, 'dashboard/etudiants/confirmer_suppression_etudiant.html', {
        'etudiant': etudiant
    })



# --- Vues pour les notes, paiements, présences (création/modification liée à un élève) ---
# Vous pouvez les créer comme des vues séparées ou des modales dans la page de détail de l'élève


@login_required
def ajouter_note(request, etudiant_id):
    """
    Ajouter une note pour un étudiant spécifique, en respectant l'école de l'utilisateur.
    Gestion de l'année scolaire active pour éviter MultipleObjectsReturned.
    """

    # 1️⃣ Récupération de l'école de l'utilisateur via profil
    ecole = getattr(getattr(request.user, 'profile', None), 'ecole', None)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # 2️⃣ Récupération de l'étudiant et vérification qu'il appartient à la même école
    etudiant = get_object_or_404(Etudiant, pk=etudiant_id, ecole=ecole)

    # 3️⃣ Récupération de l'année scolaire active pour l'école
    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()
    if not annee_active:
        messages.error(request, "Aucune année scolaire active n'a été trouvée pour votre école.")
        return redirect('liste_annees_scolaires')

    # 4️⃣ Traitement du formulaire
    if request.method == 'POST':
        form = NoteForm(request.POST, user=request.user)
        if form.is_valid():
            note = form.save(commit=False)
            note.etudiant = etudiant
            note.ecole = ecole
            note.annee_scolaire = annee_active
            note.save()
            messages.success(request, f"La note pour {etudiant.prenom} {etudiant.nom} a été ajoutée avec succès.")
            return redirect('detail_etudiant', etudiant_id=etudiant.pk)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field} : {error}")
    else:
        form = NoteForm(user=request.user)

    # 5️⃣ Rendu du template
    return render(request, 'dashboard/etudiants/ajouter_note.html', {
        'form': form,
        'etudiant': etudiant,
        'annee_active': annee_active,
    })


# ... créer modifier_note, supprimer_note, ajouter_paiement, modifier_paiement, etc. sur le même principe

# --- Génération de Certificat de Fréquentation ---



def draw_pdf_content(p, etudiant, annee_scolaire_active, settings, nb_presences, code_verification):
    """Dessine le contenu du certificat de fréquentation sur le canevas ReportLab."""
    width, height = A4
    LEFT_MARGIN = inch
    RIGHT_MARGIN = width - inch
    HEADER_Y = height - 80

    # --- 1. En-tête (Ministère & École) ---
    p.setFont("Helvetica-Bold", 12)

    # 🔹 Gestion du retour à la ligne pour le texte du ministère
    ministere_text = settings.ministere or "RÉPUBLIQUE DU MALI"
    ministere_lines = wrap(ministere_text, width=55)  # largeur max par ligne

    current_y = HEADER_Y
    for line in ministere_lines:
        p.drawRightString(RIGHT_MARGIN, current_y, line)
        current_y -= 14  # espacement entre les lignes

    # Nom de l'école
    p.setFont("Helvetica", 12)
    p.drawRightString(RIGHT_MARGIN, current_y - 10, f"École : {settings.nom_etablissement.upper()}")

    # Ligne de séparation
    p.line(LEFT_MARGIN, current_y - 40, RIGHT_MARGIN, current_y - 40)

    # --- 2. Logo ---
    if settings.logo and hasattr(settings.logo, "path"):
        try:
            p.drawImage(settings.logo.path, LEFT_MARGIN, current_y - 30, width=80, height=80, mask='auto')
        except Exception:
            p.setStrokeColorRGB(0, 0, 0)
            p.rect(LEFT_MARGIN, current_y - 30, 80, 80)
            p.setFont("Helvetica", 8)
            p.drawCentredString(LEFT_MARGIN + 40, current_y + 10, "LOGO MANQUANT")

    # --- 3. Titre Central ---
    TITLE_Y = current_y - 110
    p.setFont("Helvetica-Bold", 20)
    p.drawCentredString(width / 2, TITLE_Y, "CERTIFICAT DE FRÉQUENTATION")

    # --- 4. Corps du Texte (Détails de l'étudiant) ---
    BODY_Y = TITLE_Y - 40
    p.setFont("Helvetica", 12)

    text_lines = [
        "Nous soussignés, Direction de l'établissement, certifions que :",
        "",
        f"Matricule : {etudiant.numero_matricule or 'Non attribué'}",
        f"Nom : {etudiant.nom.upper()}",
        f"Prénom(s) : {etudiant.prenom.upper()}",
        f"Né(e) le {etudiant.date_naissance.strftime('%d/%m/%Y')} à {etudiant.lieu_naissance or 'N/A'}",
        f"Sexe : {'Masculin' if etudiant.genre == 'M' else 'Féminin'}",
        "",
        f"A fréquenté la classe de {etudiant.classe.nom_classe if etudiant.classe else 'Non assignée'}",
        f"Pendant l'Année scolaire : {annee_scolaire_active.annee}",
        "",
        f"Jours de présence enregistrés (min.) : {nb_presences} jours.",
        "",
        "En foi de quoi, le présent certificat est délivré pour servir et valoir ce que de droit."
    ]

    current_y = BODY_Y
    for line in text_lines:
        # Mettre en gras certaines informations
        if etudiant.nom.upper() in line or etudiant.prenom.upper() in line:
            p.setFont("Helvetica-Bold", 14)
        elif line.startswith("Matricule"):
            p.setFont("Helvetica-Bold", 12)
        else:
            p.setFont("Helvetica", 12)
            
        p.drawString(LEFT_MARGIN, current_y, line)
        current_y -= 18

    # --- 5. Lieu et Date de Délivrance ---
    SIGN_Y = 180
    location = (settings.adresse_etablissement.split(', ')[-1]
                if settings.adresse_etablissement else "Ville")
    
    p.setFont("Helvetica", 12)
    p.drawString(LEFT_MARGIN, SIGN_Y + 40, f"Fait à {location}, le {datetime.now().strftime('%d/%m/%Y')}")

    # --- 6. Signature, Cachet & QR Code ---
    # 6.1 QR Code
    qr = qrcode.make(f"https://monetablissement.edu.ml/verifier-certificat/{code_verification}")
    qr_io = BytesIO()
    qr.save(qr_io, format='PNG')
    qr_io.seek(0)
    
    with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_qr_file:
        tmp_qr_file.write(qr_io.read())
        qr_path = tmp_qr_file.name
    
    p.drawImage(qr_path, RIGHT_MARGIN - 100, SIGN_Y, width=80, height=80, mask='auto')
    p.setFont("Helvetica", 8)
    p.drawRightString(RIGHT_MARGIN - 60, SIGN_Y - 10, "Code de vérification")
    
    os.remove(qr_path)

    # 6.2 Signature / Cachet
    p.setFont("Helvetica-Bold", 12)
    p.drawString(RIGHT_MARGIN - 150, SIGN_Y - 40, settings.titre_signataire or "Le Directeur")

    # --- 7. Pied de Page ---
    p.setFont("Helvetica-Bold", 8)
    p.drawCentredString(width / 2, 40, f"Contact : {settings.telephone or 'N/A'} - Email : {settings.email_contact or 'N/A'}")
    p.drawCentredString(width / 2, 30, "Certificat délivré sous l’autorité du directeur.")


# -------------------------------------------------------------------------------------
# Vue Django principale
# -------------------------------------------------------------------------------------

# Le code original de la vue

# --------------------------------------------
# 1️⃣ Vue de génération automatique (PDF direct)
# --------------------------------------------
@login_required
def generer_certificat_frequentation(request, etudiant_id):
    """
    Génération du certificat de fréquentation (multi-école sécurisé)
    """

    # 🔹 Déterminer l'école de l'utilisateur connecté
    ecole_user = getattr(getattr(request.user, 'profile', None), 'ecole', None)
    if not ecole_user:
        messages.error(request, "Votre profil n'est associé à aucune école.")
        return redirect('dashboard_accueil')

    # 🔹 Récupérer uniquement un élève appartenant à cette école
    etudiant = get_object_or_404(Etudiant, pk=etudiant_id, ecole=ecole_user)
    settings = ecole_user

    # 🔹 Année scolaire active pour l’école de l’utilisateur
    annee_scolaire_active = AnneeScolaire.objects.filter(ecole=ecole_user, active=True).first()
    if not annee_scolaire_active:
        messages.error(request, "Aucune année scolaire active trouvée pour votre école.")
        return redirect('detail_etudiant', etudiant_id=etudiant.pk)

    # 🔹 Calcul des présences
    try:
        nb_presences = Presence.objects.filter(
            etudiant=etudiant,
            annee_scolaire=annee_scolaire_active,
            statut__in=['Présent', 'Retard', 'Excusé']
        ).count()
    except Exception:
        nb_presences = 0

    # ⚠️ Avertissement si peu de présences
    if nb_presences < 50:
        messages.warning(request, f"⚠️ Seulement {nb_presences} jours de présence enregistrés.")

    # 🔹 Code de vérification
    code_verification = f"CF-{etudiant.pk}-{annee_scolaire_active.annee}-{datetime.now().strftime('%H%M%S')}"

    # 🔹 Génération du PDF via ta fonction utilitaire existante
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    draw_pdf_content(p, etudiant, annee_scolaire_active, settings, nb_presences, code_verification)
    p.showPage()
    p.save()
    pdf_data = buffer.getvalue()
    buffer.close()

    # 🔹 Sauvegarde du certificat en base
    num_certificat = f"CERT-{etudiant.numero_matricule or etudiant.pk}-{annee_scolaire_active.annee.replace('-', '')}"
    pdf_filename = f"certificat_{etudiant.nom}_{etudiant.prenom}_{annee_scolaire_active.annee}.pdf"
    qr_filename = f"qr_{etudiant.nom}_{etudiant.prenom}_{annee_scolaire_active.annee}.png"

    certificat, created = CertificatFrequentation.objects.get_or_create(
        etudiant=etudiant,
        annee_scolaire=annee_scolaire_active,
        defaults={
            'date_delivrance': timezone.now().date(),
            'delivre_par': request.user,
            'numero_certificat': num_certificat,
            'ministere': settings.ministere or "",
            'academie': settings.academie or "",
            'etablissement_reference': settings.nom_etablissement or "",
            'adresse_etablissement': settings.adresse_etablissement or "",
            'mention_legale': "Certificat délivré sous l’autorité du directeur.",
            'code_verification': code_verification,
            'statut': 'valide',
            'ecole': ecole_user
        }
    )

    certificat.fichier_pdf.save(pdf_filename, ContentFile(pdf_data), save=False)

    # QR code pour vérification
    qr_content = BytesIO()
    qr_temp = qrcode.make(f"https://monetablissement.edu.ml/verifier-certificat/  {code_verification}")
    qr_temp.save(qr_content, format='PNG')
    qr_content.seek(0)
    certificat.qr_code.save(qr_filename, ContentFile(qr_content.read()), save=False)

    certificat.save()
    messages.success(request, f"✅ Certificat généré avec succès pour {etudiant.nom} {etudiant.prenom}.")
    
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{pdf_filename}"'
    return response


# -------------------------------------------------------------------------------------
# creation de certificat
# -------------------------------------------------------------------------------------
from io import BytesIO

@login_required
def creer_certificat_interface(request):
    """
    Création d'un certificat de fréquentation (PDF A4 + QR code)
    pour un élève de l'école de l'utilisateur connecté.
    """
    ecole = getattr(getattr(request.user, 'profile', None), 'ecole', None)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # 🔹 Pré-remplir depuis la fiche élève
    etudiant_id = request.GET.get('etudiant_id')
    initial_data = {}
    if etudiant_id:
        etudiant = get_object_or_404(Etudiant, pk=etudiant_id, ecole=ecole)
        initial_data['etudiant'] = etudiant
    else:
        etudiant = None

    # 🔹 Année scolaire active
    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()
    if not annee_active:
        messages.error(request, "Aucune année scolaire active trouvée pour votre école.")
        return redirect('liste_annees_scolaires')

    # 🔹 Traitement du formulaire
    if request.method == 'POST':
        form = CertificatFrequentationForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            certificat = form.save(commit=False)
            certificat.ecole = ecole
            certificat.annee_scolaire = annee_active
            certificat.date_delivrance = certificat.date_delivrance or timezone.now().date()
            certificat.delivre_par = request.user

            # --- Numéro du certificat ---
            etu = certificat.etudiant
            if not certificat.numero_certificat:
                annee_certif = certificat.date_delivrance.year
                certificat.numero_certificat = f"CERT-{etu.numero_matricule or etu.pk}-{annee_certif}"

            # --- Code de vérification & QR ---
            code_verif = f"CERT-{etu.pk}-{annee_active.annee}-{timezone.now().strftime('%H%M%S')}"
            certificat.code_verification = code_verif
            qr_url = f"https://monetablissement.edu.ml/verifier-certificat/{code_verif}"
            qr = qrcode.make(qr_url)

            qr_buffer = BytesIO()
            qr.save(qr_buffer, format='PNG')
            qr_content = qr_buffer.getvalue()
            qr_image = ImageReader(BytesIO(qr_content))
            qr_filename = f"qr_{etu.nom}_{etu.prenom}.png".replace(" ", "_")

            # --- Génération du PDF ---
            pdf_buffer = BytesIO()
            try:
                p = canvas.Canvas(pdf_buffer, pagesize=A4)
                width, height = A4
                LEFT_MARGIN, RIGHT_MARGIN = inch, width - inch
                HEADER_Y = height - 80

                # En-tête officiel
                p.setFont("Helvetica-Bold", 14)
                p.drawCentredString(width / 2, HEADER_Y, certificat.ministere)
                p.setFont("Helvetica", 12)
                p.drawCentredString(width / 2, HEADER_Y - 20, certificat.academie)
                p.drawCentredString(width / 2, HEADER_Y - 40, certificat.etablissement_reference)
                p.line(LEFT_MARGIN, HEADER_Y - 60, RIGHT_MARGIN, HEADER_Y - 60)

                # Corps du texte
                text = p.beginText(LEFT_MARGIN, HEADER_Y - 120)
                text.setFont("Helvetica", 12)
                text.textLine("Nous, soussignés, Direction de l'établissement, certifions que :")
                text.textLine("")
                text.setFont("Helvetica-Bold", 14)
                text.textLine(f"{etu.nom.upper()} {etu.prenom.upper()}")
                text.setFont("Helvetica", 12)

                date_naissance_str = etu.date_naissance.strftime('%d/%m/%Y') if etu.date_naissance else 'N/A'
                text.textLine(f"Né(e) le {date_naissance_str} à {etu.lieu_naissance or 'N/A'}")
                text.textLine(f"Sexe : {'Masculin' if etu.genre == 'M' else 'Féminin'}")
                text.textLine(f"Matricule : {etu.numero_matricule or 'Non attribué'}")
                text.textLine(f"A fréquenté la classe de {etu.classe.nom_classe if etu.classe else 'Non assignée'}.")
                text.textLine(f"Année scolaire : {annee_active.annee}")
                text.textLine("")
                text.textLine(certificat.mention_legale or 
                              "En foi de quoi, le présent certificat est délivré pour servir et valoir ce que de droit.")
                p.drawText(text)

                # QR Code
                p.drawImage(qr_image, RIGHT_MARGIN - 100, HEADER_Y - 400, width=80, height=80, mask='auto')

                # Signature et cachet
                lieu = getattr(certificat, "lieu_delivrance", None) or ecole.nom or "N/A"
                p.setFont("Helvetica", 12)
                p.drawString(LEFT_MARGIN, HEADER_Y - 420, f"Fait à {lieu}, le {timezone.now().strftime('%d/%m/%Y')}")
                p.setFont("Helvetica-Bold", 12)
                p.drawRightString(RIGHT_MARGIN - 50, HEADER_Y - 460, "Le Directeur")

                # Pied de page
                p.setFont("Helvetica-Bold", 8)
                p.drawCentredString(width / 2, 40, f"Contact : {certificat.adresse_etablissement or 'N/A'}")

                p.showPage()
                p.save()
                pdf_data = pdf_buffer.getvalue()
            finally:
                pdf_buffer.close()

            # --- Sauvegarde dans le modèle ---
            certificat.qr_code.save(qr_filename, ContentFile(qr_content), save=False)
            certificat.fichier_pdf.save(
                f"certificat_{etu.nom}_{etu.prenom}.pdf".replace(" ", "_"),
                ContentFile(pdf_data),
                save=False
            )
            certificat.statut = 'valide'
            certificat.save()

            messages.success(request, f"✅ Certificat de fréquentation généré avec succès pour {etu.nom} {etu.prenom}.")
            return redirect('detail_etudiant', pk=etu.pk)
        else:
            messages.error(request, "❌ Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = CertificatFrequentationForm(initial=initial_data, user=request.user)

    return render(request, "dashboard/etudiants/creer_certificat.html", {"form": form})

# ----------------------------------
# les views de paiement 
#------------------------------

@login_required
def liste_paiements_par_classe_etudiant(request):
    """
    Liste les paiements des étudiants regroupés par classe,
    filtrés par école et année scolaire active.
    """
    # École de l'utilisateur
    ecole = getattr(getattr(request.user, 'enseignant', None), 'ecole', None) \
            or getattr(getattr(request.user, 'etudiant', None), 'ecole', None) \
            or getattr(getattr(request.user, 'profile', None), 'ecole', None)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # Année scolaire active pour cette école
    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()
    if not annee_active:
        messages.warning(request, "Aucune année scolaire active n'est définie pour votre école.")
        return render(request, 'dashboard/paiements/paiements_par_classe_etudiant.html', {'data_par_classe': []})

    # Filtres GET
    selected_classe_id = request.GET.get('classe_filter_id')
    selected_statut = request.GET.get('statut_filter')

    try:
        selected_classe_id = int(selected_classe_id) if selected_classe_id else None
    except ValueError:
        selected_classe_id = None

    # Classes et étudiants
    classes_qs = Classe.objects.filter(ecole=ecole).order_by('nom_classe')
    etudiants_qs = Etudiant.objects.filter(
        ecole=ecole,
        annee_scolaire_inscription=annee_active
    ).select_related('classe')

    # Paiements filtrés
    paiements_qs = Paiement.objects.filter(
        etudiant__ecole=ecole,
        annee_scolaire=annee_active
    ).select_related('etudiant', 'etudiant__classe')

    # Filtrage par classe
    if selected_classe_id:
        etudiants_qs = etudiants_qs.filter(classe_id=selected_classe_id)
        paiements_qs = paiements_qs.filter(etudiant__classe_id=selected_classe_id)
        classes_qs = classes_qs.filter(id=selected_classe_id)

    # Filtrage par statut
    if selected_statut == 'payes':
        paiements_qs = paiements_qs.filter(statut='Payé')
    elif selected_statut == 'impayes_partiels':
        paiements_qs = paiements_qs.filter(statut__in=['Impayé', 'Partiel'])

    # Organisation par étudiant
    paiements_par_etudiant = {}
    for paiement in paiements_qs:
        etudiant_id = paiement.etudiant_id
        if etudiant_id not in paiements_par_etudiant:
            paiements_par_etudiant[etudiant_id] = {'payes': [], 'impayes_partiels': []}
        key = 'payes' if paiement.statut == 'Payé' else 'impayes_partiels'
        paiements_par_etudiant[etudiant_id][key].append(paiement)

    # Organisation par classe
    data_par_classe = []
    for classe in classes_qs:
        etudiants_data = []
        for etudiant in etudiants_qs.filter(classe=classe):
            paiements_etudiant = paiements_par_etudiant.get(etudiant.id, {'payes': [], 'impayes_partiels': []})
            paiements_list = paiements_etudiant['payes'] + paiements_etudiant['impayes_partiels']
            if paiements_list:
                etudiants_data.append({
                    'etudiant': etudiant,
                    'total_du': sum(p.montant_du for p in paiements_list),
                    'total_paye': sum(p.montant for p in paiements_list),
                    'paiements_payes': paiements_etudiant['payes'],
                    'paiements_impayes_partiels': paiements_etudiant['impayes_partiels'],
                })
        if etudiants_data:
            data_par_classe.append({'classe': classe, 'etudiants_data': etudiants_data})

    context = {
        'annee_active': annee_active,
        'data_par_classe': data_par_classe,
        'toutes_les_classes': Classe.objects.filter(ecole=ecole).order_by('nom_classe'),
        'selected_classe_id': selected_classe_id,
        'selected_statut': selected_statut,
    }
    return render(request, 'dashboard/paiements/paiements_par_classe_etudiant.html', context)





@login_required
def liste_paiements_impayes(request):
    ecole = getattr(getattr(request.user, 'enseignant', None), 'ecole', None) \
            or getattr(getattr(request.user, 'profile', None), 'ecole', None)
    if not ecole:
        messages.warning(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()
    paiements_impayes = []
    if annee_active:
        paiements_impayes = Paiement.objects.filter(
            etudiant__ecole=ecole,
            annee_scolaire=annee_active,
            statut__in=['Impayé', 'Partiel']
        ).select_related('etudiant', 'etudiant__classe').order_by('etudiant__classe__nom_classe', 'etudiant__nom')

    return render(request, 'dashboard/paiements/paiements_impayes.html', {
        'paiements_impayes': paiements_impayes,
        'annee_active': annee_active,
    })

@login_required
def liste_paiements_payes(request):
    ecole = getattr(getattr(request.user, 'enseignant', None), 'ecole', None) \
            or getattr(getattr(request.user, 'profile', None), 'ecole', None)
    if not ecole:
        messages.warning(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()
    paiements_payes = []
    if annee_active:
        paiements_payes = Paiement.objects.filter(
            etudiant__ecole=ecole,
            annee_scolaire=annee_active,
            statut='Payé'
        ).select_related('etudiant', 'etudiant__classe').order_by('etudiant__classe__nom_classe', '-date_paiement')

    return render(request, 'dashboard/paiements/paiements_payes.html', {
        'paiements_payes': paiements_payes,
        'annee_active': annee_active,
    })







@login_required
def generer_recu_paiement(request, paiement_id):
    """
    Génère un reçu PDF pour un paiement spécifique.
    Utilise wkhtmltopdf si disponible, sinon WeasyPrint comme solution de secours.
    Inclut automatiquement le logo de l’école dans le PDF.
    """

    # --- 1️⃣ Vérification du paiement ---
    try:
        paiement = get_object_or_404(
            Paiement,
            pk=paiement_id,
            etudiant__ecole=request.user.profile.ecole
        )
    except AttributeError:
        # Cas des superadmins sans profil.ecole
        paiement = get_object_or_404(Paiement, pk=paiement_id)

    etudiant = paiement.etudiant
    ecole = paiement.etudiant.ecole or EcoleSettings.objects.filter(unique_instance=True).first()

    # --- 2️⃣ Préparation du logo (chemin absolu pour PDF) ---
    logo_url = ""
    if ecole and ecole.logo:
        try:
            logo_url = request.build_absolute_uri(ecole.logo.url)
        except Exception:
            logo_url = ""  # au cas où MEDIA_URL n’est pas dispo

    # --- 3️⃣ Contexte du reçu ---
    context = {
        'paiement': paiement,
        'etudiant': etudiant,
        'ecole': ecole,
        'logo_url': logo_url,  # ajouté pour affichage dans le template
        
    }

    # --- 4️⃣ Rendu HTML ---
    html_string = render_to_string('dashboard/paiements/recu_paiement.html', context)

    # --- 5️⃣ Options wkhtmltopdf ---
    options = {
        'page-size': 'A5',
        'encoding': "UTF-8",
        'no-outline': None,
        'margin-top': '0.4in',
        'margin-bottom': '0.4in',
        'margin-left': '0.3in',
        'margin-right': '0.3in',
    }

    pdf_data = None

    # --- 6️⃣ Génération du PDF ---
    try:
        # Cherche wkhtmltopdf sur le système
        path_wkhtmltopdf = '/usr/bin/wkhtmltopdf'
        if not os.path.exists(path_wkhtmltopdf):
            path_wkhtmltopdf = shutil.which("wkhtmltopdf")

        if path_wkhtmltopdf:
            config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
            pdf_data = pdfkit.from_string(html_string, False, options=options, configuration=config)
        else:
            raise OSError("wkhtmltopdf non trouvé")

    except Exception as e:
        # --- 7️⃣ Fallback WeasyPrint ---
        messages.warning(request, f"wkhtmltopdf introuvable — génération via WeasyPrint ({e})")
        pdf_data = HTML(string=html_string, base_url=request.build_absolute_uri("/")).write_pdf()

    # --- 8️⃣ Réponse HTTP ---
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Recu_{etudiant.nom}_{paiement.id}.pdf"'
    return response






# --- CRUD pour les Années Scolaires ---



@login_required
def liste_annees_scolaires(request):
    """
    Liste toutes les années scolaires pour l'école de l'utilisateur connecté
    et calcule l'année scolaire actuelle selon le calendrier.
    """
    # 🔹 Récupérer l'école de l'utilisateur
    ecole = (
        getattr(getattr(request.user, 'enseignant', None), 'ecole', None)
        or getattr(getattr(request.user, 'etudiant', None), 'ecole', None)
        or getattr(getattr(request.user, 'profile', None), 'ecole', None)
        or (EcoleSettings.objects.first() if request.user.is_superuser else None)
    )

    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # 🔹 Filtrer par école
    annees_scolaires = AnneeScolaire.objects.filter(ecole=ecole).order_by('-annee')

    # 🔹 Calcul de l'année scolaire actuelle selon le calendrier
    today = datetime.now()
    if today.month >= 8:
        start_year = today.year
        end_year = today.year + 1
    else:
        start_year = today.year - 1
        end_year = today.year
    current_academic_year = f"{start_year}-{end_year}"

    context = {
        'annees_scolaires': annees_scolaires,
        'current_calendar_year': today.year,
        'current_academic_year_string': current_academic_year,
        'ecole': ecole,
    }
    return render(request, 'dashboard/annees_scolaires/liste_annees_scolaires.html', context)



@login_required
def creer_annee_scolaire(request):
    """Créer une nouvelle année scolaire pour l'école de l'utilisateur"""
    ecole = get_user_ecole(request)  # Récupération sécurisée de l'école
    if request.method == 'POST':
        form = AnneeScolaireForm(request.POST, ecole=ecole)
        if form.is_valid():
            annee = form.save(commit=False)
            annee.ecole = ecole  # Assigner l'école avant la sauvegarde
            annee.save()
            messages.success(request, "Année scolaire ajoutée avec succès.")
            return redirect('liste_annees_scolaires')
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = AnneeScolaireForm(ecole=ecole)

    return render(request, 'dashboard/annees_scolaires/form_annee_scolaire.html', {
        'form': form,
        'action': 'Créer'
    })


@login_required
def modifier_annee_scolaire(request, pk):
    """Modifier une année scolaire existante de l'école de l'utilisateur"""
    ecole = get_user_ecole(request)
    annee = get_object_or_404(AnneeScolaire, pk=pk, ecole=ecole)  # ⚠️ Filtrage par école pour sécurité

    if request.method == 'POST':
        form = AnneeScolaireForm(request.POST, instance=annee, ecole=ecole)
        if form.is_valid():
            form.save()
            messages.success(request, "Année scolaire mise à jour avec succès.")
            return redirect('liste_annees_scolaires')
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = AnneeScolaireForm(instance=annee, ecole=ecole)

    return render(request, 'dashboard/annees_scolaires/form_annee_scolaire.html', {
        'form': form,
        'action': 'Modifier'
    })



@login_required
def supprimer_annee_scolaire(request, pk):
    """Supprimer une année scolaire après confirmation"""
    annee = get_object_or_404(AnneeScolaire, pk=pk)
    
    if request.method == 'POST':
        annee_nom = annee.annee
        annee.delete()
        messages.success(request, f"L'année scolaire {annee_nom} a été supprimée.")
        return redirect('liste_annees_scolaires')
    
    return render(request, 'dashboard/annees_scolaires/confirmer_suppression_annee_scolaire.html', {
        'annee': annee
    })


# ==========================
# CRUD Classes
# ==========================

def get_user_ecole(request):
    """Récupère l'école de l'utilisateur connecté"""
    return (
        getattr(getattr(request.user, 'enseignant', None), 'ecole', None)
        or getattr(getattr(request.user, 'etudiant', None), 'ecole', None)
        or getattr(getattr(request.user, 'profile', None), 'ecole', None)
        or (EcoleSettings.objects.first() if request.user.is_superuser else None)
    )


@login_required
def liste_classes(request):
    """
    Liste toutes les classes de l'école de l'utilisateur,
    triées par année scolaire et nom, avec filtrage strict par école.
    """
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    classes = Classe.objects.filter(ecole=ecole) \
                           .select_related('annee_scolaire', 'enseignant_principal') \
                           .order_by('annee_scolaire__annee', 'nom_classe')

    return render(request, 'dashboard/classes/liste_classes.html', {
        'classes': classes,
        'ecole': ecole
    })


@login_required
def creer_classe(request):
    """
    Créer une nouvelle classe pour l'école de l'utilisateur
    et assigner automatiquement l'année scolaire active.
    Filtre les champs du formulaire selon l'école.
    """
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # Récupérer la première année active pour cette école
    annee_active = AnneeScolaire.objects.filter(ecole=ecole, active=True).order_by('-annee').first()
    if not annee_active:
        messages.warning(request, "Veuillez définir une année scolaire active avant de créer une classe.")
        return redirect('liste_annees_scolaires')

    if request.method == 'POST':
        form = ClasseForm(request.POST, ecole=ecole)
        if form.is_valid():
            classe = form.save(commit=False)
            classe.ecole = ecole
            classe.annee_scolaire = annee_active  # assigner automatiquement l'année active
            classe.save()
            messages.success(request, f"Classe '{classe.nom_classe}' ajoutée avec succès.")
            return redirect('liste_classes')
        messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = ClasseForm(ecole=ecole)

    return render(request, 'dashboard/classes/form_classe.html', {
        'form': form,
        'action': 'Créer',
        'annee_active': annee_active,
    })

@login_required
def modifier_classe(request, pk):
    """
    Modifier une classe existante uniquement si elle appartient à l'école de l'utilisateur.
    Passe l'école au formulaire pour filtrer les champs liés (ex: matières, enseignants).
    """
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    classe = get_object_or_404(Classe, pk=pk, ecole=ecole)

    if request.method == 'POST':
        form = ClasseForm(request.POST, instance=classe, ecole=ecole)
        if form.is_valid():
            form.save()
            messages.success(request, "Classe mise à jour avec succès.")
            return redirect('liste_classes')
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = ClasseForm(instance=classe, ecole=ecole)

    return render(request, 'dashboard/classes/form_classe.html', {
        'form': form,
        'action': 'Modifier'
    })


@login_required
def supprimer_classe(request, pk):
    """
    Supprimer une classe uniquement si elle appartient à l'école de l'utilisateur.
    """
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    classe = get_object_or_404(Classe, pk=pk, ecole=ecole)

    if request.method == 'POST':
        nom_classe = classe.nom_classe
        classe.delete()
        messages.success(request, f"La classe {nom_classe} a été supprimée.")
        return redirect('liste_classes')

    return render(request, 'dashboard/classes/confirmer_suppression_classe.html', {
        'classe': classe
    })


# ------------------ MATIERES ------------------

@login_required
def liste_matieres(request):
    """Liste des matières filtrée par école"""
    ecole = getattr(getattr(request.user, 'enseignant', None), 'ecole', None) \
            or getattr(getattr(request.user, 'profile', None), 'ecole', None)

    matieres = Matiere.objects.filter(ecole=ecole).order_by('nom')
    return render(request, 'dashboard/matieres/liste_matieres.html', {'matieres': matieres, 'ecole': ecole})


@login_required
def creer_matiere(request):
    """Créer une matière liée à l'école de l'utilisateur"""
    ecole = getattr(getattr(request.user, 'enseignant', None), 'ecole', None) \
            or getattr(getattr(request.user, 'profile', None), 'ecole', None)

    if request.method == 'POST':
        form = MatiereForm(request.POST)
        if form.is_valid():
            matiere = form.save(commit=False)
            matiere.ecole = ecole
            matiere.save()
            messages.success(request, "Matière ajoutée avec succès.")
            return redirect('liste_matieres')
        messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = MatiereForm()
    
    return render(request, 'dashboard/matieres/form_matiere.html', {'form': form, 'action': 'Créer'})


@login_required
def modifier_matiere(request, pk):
    """Modifier une matière uniquement si elle appartient à l'école de l'utilisateur"""
    ecole = getattr(getattr(request.user, 'enseignant', None), 'ecole', None) \
            or getattr(getattr(request.user, 'profile', None), 'ecole', None)

    matiere = get_object_or_404(Matiere, pk=pk, ecole=ecole)

    if request.method == 'POST':
        form = MatiereForm(request.POST, instance=matiere)
        if form.is_valid():
            form.save()
            messages.success(request, "Matière mise à jour avec succès.")
            return redirect('liste_matieres')
        messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = MatiereForm(instance=matiere)
    
    return render(request, 'dashboard/matieres/form_matiere.html', {'form': form, 'action': 'Modifier'})


@login_required
def supprimer_matiere(request, pk):
    """Supprimer une matière uniquement si elle appartient à l'école de l'utilisateur"""
    ecole = getattr(getattr(request.user, 'enseignant', None), 'ecole', None) \
            or getattr(getattr(request.user, 'profile', None), 'ecole', None)

    matiere = get_object_or_404(Matiere, pk=pk, ecole=ecole)

    if request.method == 'POST':
        nom_matiere = matiere.nom
        matiere.delete()
        messages.success(request, f"La matière {nom_matiere} a été supprimée.")
        return redirect('liste_matieres')
    
    return render(request, 'dashboard/matieres/confirmer_suppression_matiere.html', {'matiere': matiere})


# ==========================
# CRUD Enseignants
# ==========================

@login_required
def liste_enseignants(request):
    """Liste des enseignants filtrée par l'école de l'utilisateur"""
    ecole = (
        getattr(getattr(request.user, 'enseignant', None), 'ecole', None)
        or getattr(getattr(request.user, 'profile', None), 'ecole', None)
        or (EcoleSettings.objects.first() if request.user.is_superuser else None)
    )

    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    enseignants = Enseignant.objects.filter(ecole=ecole).order_by('nom', 'prenom')
    return render(request, 'dashboard/enseignants/liste_enseignants.html', {'enseignants': enseignants, 'ecole': ecole})


@login_required
def creer_enseignant(request):
    ecole = (
        getattr(getattr(request.user, 'enseignant', None), 'ecole', None)
        or getattr(getattr(request.user, 'profile', None), 'ecole', None)
        or (EcoleSettings.objects.first() if request.user.is_superuser else None)
    )

    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    if request.method == 'POST':
        form = EnseignantForm(request.POST)
        if form.is_valid():
            enseignant = form.save(commit=False)
            enseignant.ecole = ecole  # Associer automatiquement l'école
            enseignant.save()
            messages.success(request, f"L'enseignant {enseignant.prenom} {enseignant.nom} a été ajouté avec succès.")
            return redirect('liste_enseignants')
        messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = EnseignantForm()
    
    return render(request, 'dashboard/enseignants/form_enseignant.html', {'form': form, 'action': 'Créer'})


@login_required
def modifier_enseignant(request, pk):
    enseignant = get_object_or_404(Enseignant, pk=pk)

    # Vérifier que l'enseignant appartient à l'école de l'utilisateur
    ecole = (
        getattr(getattr(request.user, 'enseignant', None), 'ecole', None)
        or getattr(getattr(request.user, 'profile', None), 'ecole', None)
        or (EcoleSettings.objects.first() if request.user.is_superuser else None)
    )
    if not request.user.is_superuser and enseignant.ecole != ecole:
        messages.error(request, "Vous n'êtes pas autorisé à modifier cet enseignant.")
        return redirect('dashboard_accueil')

    if request.method == 'POST':
        form = EnseignantForm(request.POST, instance=enseignant)
        if form.is_valid():
            form.save()
            messages.success(request, f"L'enseignant {enseignant.prenom} {enseignant.nom} a été mis à jour.")
            return redirect('liste_enseignants')
        messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = EnseignantForm(instance=enseignant)
    
    return render(request, 'dashboard/enseignants/form_enseignant.html', {'form': form, 'action': 'Modifier'})


@login_required
def supprimer_enseignant(request, pk):
    enseignant = get_object_or_404(Enseignant, pk=pk)

    ecole = (
        getattr(getattr(request.user, 'enseignant', None), 'ecole', None)
        or getattr(getattr(request.user, 'profile', None), 'ecole', None)
        or (EcoleSettings.objects.first() if request.user.is_superuser else None)
    )
    if not request.user.is_superuser and enseignant.ecole != ecole:
        messages.error(request, "Vous n'êtes pas autorisé à supprimer cet enseignant.")
        return redirect('dashboard_accueil')

    if request.method == 'POST':
        nom_complet = f"{enseignant.prenom} {enseignant.nom}"
        enseignant.delete()
        messages.success(request, f"L'enseignant {nom_complet} a été supprimé.")
        return redirect('liste_enseignants')
    
    return render(request, 'dashboard/enseignants/confirmer_suppression_enseignant.html', {'enseignant': enseignant})


# ==========================
# CRUD Programmes Matières
# ==========================
@login_required
def liste_programmes_matiere(request):
    # Récupérer l'école de l'utilisateur
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # Récupérer l'année scolaire active pour cette école
    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()
    if not annee_active:
        messages.warning(request, "Aucune année scolaire active n'est définie pour votre école.")
        return redirect('liste_annees_scolaires')

    # Filtrer les programmes de matière par école et année scolaire active
    programmes = ProgrammeMatiere.objects.filter(
        classe__ecole=ecole,
        classe__annee_scolaire=annee_active
    ).select_related('classe', 'matiere', 'enseignant').order_by('classe__nom_classe', 'matiere__nom')

    return render(request, 'dashboard/programmes_matiere/liste_programmes_matiere.html', {
        'programmes': programmes,
        'annee_active': annee_active,
    })


@login_required
def creer_programme_matiere(request):
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    if request.method == 'POST':
        form = ProgrammeMatiereForm(request.POST, ecole=ecole)
        if form.is_valid():
            programme = form.save(commit=False)
            programme.ecole = ecole  # affecte l’école automatiquement
            programme.save()
            messages.success(request, "Programme matière ajouté avec succès.")
            return redirect('liste_programmes_matiere')
        messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = ProgrammeMatiereForm(ecole=ecole)

    return render(request, 'dashboard/programmes_matiere/form_programme_matiere.html', {
        'form': form,
        'action': 'Créer'
    })


@login_required
def modifier_programme_matiere(request, pk):
    # Récupérer l'école de l'utilisateur connecté
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # Récupérer le programme matière pour l'école courante
    programme = get_object_or_404(
        ProgrammeMatiere.objects.select_related('classe', 'matiere', 'enseignant'),
        pk=pk,
        classe__ecole=ecole  # 🔥 Filtrage strict par école via la classe
    )

    # Création du formulaire
    if request.method == 'POST':
        form = ProgrammeMatiereForm(request.POST, instance=programme, ecole=ecole)
        if form.is_valid():
            form.save()
            messages.success(request, "Programme matière mis à jour avec succès.")
            return redirect('liste_programmes_matiere')
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = ProgrammeMatiereForm(instance=programme, ecole=ecole)

    return render(request, 'dashboard/programmes_matiere/form_programme_matiere.html', {
        'form': form,
        'action': 'Modifier',
        'programme': programme
    })



@login_required
def supprimer_programme_matiere(request, pk):
    # récupérer l'école de l'utilisateur
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # essayer de récupérer le programme pour cette école
    try:
        # on suppose que ProgrammeMatiere est lié à une classe (classe) et que la classe a un champ ecole
        programme = ProgrammeMatiere.objects.get(pk=pk, classe__ecole=ecole)
    except ProgrammeMatiere.DoesNotExist:
        # si superuser, on peut permettre l'accès sans filtre (optionnel)
        if request.user.is_superuser:
            try:
                programme = ProgrammeMatiere.objects.get(pk=pk)
            except ProgrammeMatiere.DoesNotExist:
                messages.error(request, "Le programme demandé n'existe pas.")
                return redirect('liste_programmes_matiere')
        else:
            messages.error(request, "Programme introuvable ou n'appartient pas à votre école.")
            return redirect('liste_programmes_matiere')

    # confirmation POST -> suppression
    if request.method == 'POST':
        nom_complet = f"{programme.matiere.nom} pour {programme.classe.nom_classe}"
        programme.delete()
        messages.success(request, f"Le programme matière '{nom_complet}' a été supprimé.")
        return redirect('liste_programmes_matiere')

    # GET -> afficher confirmation
    return render(request, 'dashboard/programmes_matiere/confirmer_suppression_programme_matiere.html', {
        'programme': programme
    })


# --- CRUD pour les Notes ---
# Les fonctions ajouter/modifier/supprimer note sont généralement liées à la page détail élève.
# On a déjà ajouter_note ci-dessus. Voici un exemple pour modifier_note.
@login_required
def modifier_note(request, pk):
    # Récupérer l'école de l'utilisateur connecté
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # Récupérer la note si elle appartient à un étudiant de cette école
    note = get_object_or_404(
        Note.objects.select_related('etudiant', 'etudiant__classe'),
        pk=pk,
        etudiant__classe__ecole=ecole  # 🔥 Filtrage par école
    )
    etudiant_id = note.etudiant.pk

    if request.method == 'POST':
        form = NoteForm(request.POST, instance=note)
        if form.is_valid():
            form.save()
            messages.success(request, "Note modifiée avec succès.")
            return redirect('detail_etudiant', etudiant_id=etudiant_id)
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = NoteForm(instance=note)

    return render(request, 'dashboard/etudiants/modifier_note.html', {
        'form': form,
        'note': note
    })


@login_required
def supprimer_note(request, pk):
    # Récupérer l'école de l'utilisateur connecté
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # Récupérer la note uniquement si elle appartient à un étudiant de cette école
    note = get_object_or_404(
        Note.objects.select_related('etudiant', 'etudiant__classe'),
        pk=pk,
        etudiant__classe__ecole=ecole  # 🔥 Filtrage strict
    )
    etudiant_id = note.etudiant.pk

    if request.method == 'POST':
        note.delete()
        messages.success(request, "Note supprimée avec succès.")
        return redirect('detail_etudiant', etudiant_id=etudiant_id)

    return render(request, 'dashboard/etudiants/confirmer_suppression_note.html', {
        'note': note
    })



# ==========================
# CRUD Paiements
# ==========================
@login_required
def ajouter_paiement(request, etudiant_id):
    # Récupérer l'école de l'utilisateur connecté
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # Vérifier que l'étudiant appartient à la même école
    etudiant = get_object_or_404(Etudiant, pk=etudiant_id, ecole=ecole)

    # Récupérer l'année scolaire active pour cette école
    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()

    if request.method == 'POST':
        form = PaiementForm(request.POST, request.FILES, ecole=ecole, initial={'etudiant': etudiant})
        if form.is_valid():
            paiement = form.save(commit=False)
            paiement.etudiant = etudiant
            paiement.enregistre_par = request.user
            paiement.ecole = ecole
            paiement.annee_scolaire = annee_active
            paiement.save()
            messages.success(request, f"Paiement ajouté avec succès pour {etudiant.prenom} {etudiant.nom}.")
            return redirect('detail_etudiant', etudiant_id=etudiant.pk)
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = PaiementForm(ecole=ecole, initial={'etudiant': etudiant, 'annee_scolaire': annee_active})

    return render(
        request,
        'dashboard/etudiants/ajouter_paiement.html',
        {'form': form, 'etudiant': etudiant}
    )

    

@login_required
def modifier_paiement(request, pk):
    # Récupération de l'école de l'utilisateur connecté
    user_ecole = get_user_ecole(request)
    if not user_ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # Récupération du paiement appartenant à un étudiant de cette école
    paiement = get_object_or_404(
        Paiement.objects.select_related('etudiant', 'etudiant__classe'),
        pk=pk,
        etudiant__classe__ecole=user_ecole  # 🔥 Filtrage par école
    )
    etudiant = paiement.etudiant
    ecole = etudiant.classe.ecole  # L'école de l'étudiant

    # Vérification de permission : superuser ou école correspondante
    if not request.user.is_superuser and user_ecole != ecole:
        messages.error(request, "Vous n'êtes pas autorisé à modifier ce paiement.")
        return redirect('dashboard_accueil')

    # Année scolaire active pour l'école
    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()

    if request.method == 'POST':
        form = PaiementForm(request.POST, request.FILES, instance=paiement, ecole=ecole)
        if form.is_valid():
            paiement_modifie = form.save(commit=False)
            paiement_modifie.annee_scolaire = annee_active
            paiement_modifie.save()
            messages.success(request, "Paiement modifié avec succès.")
            return redirect('detail_etudiant', etudiant_id=etudiant.pk)
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = PaiementForm(instance=paiement, ecole=ecole)

    return render(
        request,
        'dashboard/etudiants/modifier_paiement.html',
        {'form': form, 'paiement': paiement, 'etudiant': etudiant}
    )


@login_required
def supprimer_paiement(request, pk):
    # Récupération de l'école de l'utilisateur connecté
    user_ecole = get_user_ecole(request)
    if not user_ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # Récupération du paiement appartenant à l'école de l'utilisateur
    paiement = get_object_or_404(
        Paiement.objects.select_related('etudiant', 'etudiant__classe'),
        pk=pk,
        etudiant__classe__ecole=user_ecole  # 🔥 Filtrage strict
    )
    etudiant_id = paiement.etudiant.pk

    # Vérification de permission : superuser ou école correspondante
    if not request.user.is_superuser and paiement.etudiant.classe.ecole != user_ecole:
        messages.error(request, "Vous n'êtes pas autorisé à supprimer ce paiement.")
        return redirect('dashboard_accueil')

    if request.method == 'POST':
        paiement.delete()
        messages.success(request, "Paiement supprimé avec succès.")
        return redirect('detail_etudiant', etudiant_id=etudiant_id)

    return render(
        request,
        'dashboard/etudiants/confirmer_suppression_paiement.html',
        {'paiement': paiement}
    )





@login_required
def liste_paiements_impayes(request):
    # Récupérer l'école de l'utilisateur
    ecole = getattr(getattr(request.user, 'enseignant', None), 'ecole', None) \
            or getattr(getattr(request.user, 'profile', None), 'ecole', None)

    if not ecole:
        messages.warning(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # Récupérer l'année scolaire active pour cette école
    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()
    if not annee_active:
        messages.warning(request, "Aucune année scolaire active n'est définie pour votre école.")
        return redirect('liste_annees_scolaires')

    # Paiements impayés ou partiels pour les élèves de l'école
    paiements_impayes = Paiement.objects.filter(
        etudiant__ecole=ecole,
        annee_scolaire=annee_active,
        statut__in=['Impayé', 'Partiel']
    ).select_related('etudiant', 'etudiant__classe').order_by(
        'etudiant__classe__nom_classe', 'etudiant__nom', 'etudiant__prenom'
    )

    context = {
        'paiements_impayes': paiements_impayes,
        'annee_active': annee_active,
        'ecole': ecole,
    }
    return render(request, 'dashboard/paiements/liste_paiements_impayes.html', context)



# ==========================
# Marquer les présences pour une classe
# ==========================
@login_required
def marquer_presence_classe(request, classe_id):
    # Récupérer l'école de l'utilisateur
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    matiere_id = request.GET.get('matiere_id')
    classe = get_object_or_404(Classe, pk=classe_id, ecole=ecole)
    date_aujourdhui = date.today()

    # Année scolaire active pour cette école
    annee_active = AnneeScolaire.objects.filter(ecole=ecole, active=True).first()
    if not annee_active:
        messages.warning(request, "Veuillez définir une année scolaire active avant de marquer les présences.")
        return redirect('liste_annees_scolaires')

    # Étudiants inscrits pour l'année active et dans cette école
    etudiants = Etudiant.objects.filter(
        classe=classe,
        annee_scolaire_inscription=annee_active,
        ecole=ecole
    ).order_by('nom', 'prenom')

    # Formset
    PresenceFormSet = formset_factory(PresenceForm, extra=0)
    initial_data = []

    # Présences déjà enregistrées pour cette classe, école et date
    existing_presences = {
        p.etudiant_id: p for p in Presence.objects.filter(
            classe=classe,
            date=date_aujourdhui,
            annee_scolaire=annee_active,
            ecole=ecole
        )
    }

    for etudiant in etudiants:
        p = existing_presences.get(etudiant.id)
        if p:
            initial_data.append({
                'id': p.id,
                'etudiant': etudiant,
                'etudiant_obj': etudiant,
                'classe': classe,
                'date': date_aujourdhui,
                'annee_scolaire': annee_active,
                'statut': p.statut,
                'matiere': p.matiere.pk if p.matiere else None,
                'heure_debut_cours': p.heure_debut_cours,
                'heure_fin_cours': p.heure_fin_cours,
                'motif_absence_retard': p.motif_absence_retard,
                'justificatif_fourni': p.justificatif_fourni,
                'est_present': p.statut == 'Présent',
                'statut_detail': p.statut if p.statut != 'Présent' else '',
            })
        else:
            initial_data.append({
                'etudiant': etudiant,
                'etudiant_obj': etudiant,
                'classe': classe,
                'date': date_aujourdhui,
                'annee_scolaire': annee_active,
                'statut': 'Présent',
                'est_present': True,
                'statut_detail': '',
            })

    if request.method == 'POST':
        formset = PresenceFormSet(request.POST, initial=initial_data)
        if formset.is_valid():
            try:
                with transaction.atomic():
                    for form in formset:
                        est_present = form.cleaned_data.get('est_present')
                        statut_detail = form.cleaned_data.get('statut_detail')
                        presence, created = Presence.objects.get_or_create(
                            etudiant=form.cleaned_data['etudiant'],
                            classe=form.cleaned_data['classe'],
                            date=form.cleaned_data['date'],
                            annee_scolaire=form.cleaned_data['annee_scolaire'],
                            ecole=ecole,  # ✅ Filtre par école
                            defaults={'statut': 'Présent'}
                        )
                        if est_present:
                            presence.statut = 'Présent'
                            presence.matiere = None
                            presence.heure_debut_cours = None
                            presence.heure_fin_cours = None
                            presence.motif_absence_retard = ''
                            presence.justificatif_fourni = False
                        else:
                            presence.statut = statut_detail or 'Absent'
                            presence.matiere = form.cleaned_data.get('matiere')
                            presence.heure_debut_cours = form.cleaned_data.get('heure_debut_cours')
                            presence.heure_fin_cours = form.cleaned_data.get('heure_fin_cours')
                            presence.motif_absence_retard = form.cleaned_data.get('motif_absence_retard', '')
                            presence.justificatif_fourni = form.cleaned_data.get('justificatif_fourni', False)
                        presence.save()
                messages.success(request, f"Présences enregistrées pour la classe {classe.nom_classe}.")
                return redirect('liste_classes')
            except Exception as e:
                messages.error(request, f"Erreur lors de l'enregistrement : {e}")
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        formset = PresenceFormSet(initial=initial_data)

    context = {
        'classe': classe,
        'date_aujourdhui': date_aujourdhui,
        'formset': formset,
        'annee_active': annee_active,
        'toutes_les_classes': Classe.objects.filter(ecole=ecole).order_by('nom_classe'),
        'toutes_les_matieres': Matiere.objects.filter(ecole=ecole).order_by('nom'),
        'matiere_id': matiere_id,
    }
    return render(request, 'dashboard/presences/marquer_presence_classe.html', context)



@login_required
def liste_presences(request):
    """
    Vue pour afficher la liste de toutes les présences d'une école.
    Permet un filtrage par classe et par date.
    """

    # 🔹 1️⃣ Récupérer l'école de l'utilisateur
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # 🔹 2️⃣ Liste de toutes les classes pour filtre
    classes = Classe.objects.filter(ecole=ecole).order_by('nom_classe')

    # 🔹 3️⃣ Récupération de filtres depuis GET
    classe_id = request.GET.get('classe')
    date_filtre = request.GET.get('date')

    classe = None
    if classe_id:
        try:
            classe = Classe.objects.get(pk=classe_id, ecole=ecole)
        except Classe.DoesNotExist:
            messages.warning(request, "La classe sélectionnée n'existe pas.")
            classe = None

    # Filtrage par année scolaire active
    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()
    if not annee_active:
        messages.warning(request, "Aucune année scolaire active définie pour votre école.")
        return redirect('dashboard_accueil')

    # 🔹 4️⃣ Construction de la queryset des présences
    presences = Presence.objects.filter(ecole=ecole, annee_scolaire=annee_active)

    if classe:
        presences = presences.filter(classe=classe)

    if date_filtre:
        try:
            date_obj = datetime.strptime(date_filtre, '%Y-%m-%d').date()
            presences = presences.filter(date=date_obj)
        except ValueError:
            messages.warning(request, "La date sélectionnée est invalide.")

    presences = presences.select_related('etudiant', 'classe').order_by('date', 'etudiant__nom')

    # 🔹 5️⃣ Contexte pour le template
    context = {
        'presences': presences,
        'classes': classes,
        'classe_selectionnee': classe,
        'date_filtre': date_filtre,
        'annee_active': annee_active,
    }

    return render(request, 'dashboard/presences/liste_presences.html', context)




# ==========================
# Suivi des présences
# ==========================
@login_required
def suivi_presence_classe(request, classe_id):
    """
    Vue de suivi des présences d'une classe, filtrée strictement par l'école de l'utilisateur connecté.
    """
    # 🔹 Récupération de l'école de l'utilisateur
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # 🔹 Classe filtrée par l'école de l'utilisateur
    classe = get_object_or_404(Classe, pk=classe_id, ecole=ecole)

    # 🔹 Année scolaire active pour cette école
    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()
    if not annee_active:
        messages.warning(request, "Aucune année scolaire active n'est définie pour votre école.")
        return redirect('liste_annees_scolaires')

    # 🔹 Filtre par date optionnel
    date_filtre = request.GET.get('date')
    if date_filtre:
        try:
            date_filtre = datetime.strptime(date_filtre, '%Y-%m-%d').date()
        except ValueError:
            date_filtre = None

    # 🔹 Étudiants inscrits dans cette classe et cette école pour l'année active
    etudiants = Etudiant.objects.filter(
        classe=classe,
        annee_scolaire_inscription=annee_active,
        ecole=ecole
    ).order_by('nom', 'prenom')

    # 🔹 Présences par étudiant
    presences_par_etudiant = {}
    for etudiant in etudiants:
        presences_query = Presence.objects.filter(
            etudiant=etudiant,
            classe=classe,
            annee_scolaire=annee_active,
            ecole=ecole
        )
        if date_filtre:
            presences_query = presences_query.filter(date=date_filtre)
        presences_par_etudiant[etudiant] = presences_query.order_by('-date')

    # 🔹 Contexte
    context = {
        'classe': classe,
        'presences_par_etudiant': presences_par_etudiant,
        'date_filtre': date_filtre,
        'annee_active': annee_active,
        'toutes_les_classes': Classe.objects.filter(ecole=ecole).order_by('nom_classe'),
    }

    return render(request, 'dashboard/presences/suivi_presence_classe.html', context)
@login_required
def suivi_presence_eleve(request, etudiant_id):
    """
    Suivi des présences d'un élève, filtré strictement par l'école de l'utilisateur connecté.
    """
    # 🔹 Récupération de l'école de l'utilisateur
    ecole_utilisateur = get_user_ecole(request)
    if not ecole_utilisateur:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # 🔹 Vérification : l'élève appartient bien à l'école de l'utilisateur
    etudiant = get_object_or_404(
        Etudiant.objects.select_related('classe', 'classe__ecole'),
        pk=etudiant_id,
        ecole=ecole_utilisateur
    )

    # 🔹 Année scolaire active pour cette école
    annee_active = AnneeScolaire.objects.filter(
        ecole=ecole_utilisateur,
        active=True
    ).first()

    if not annee_active:
        messages.warning(request, "Aucune année scolaire active n'est définie pour votre école.")
        return redirect('liste_annees_scolaires')

    # 🔹 Présences filtrées par élève, école et année scolaire
    presences = Presence.objects.filter(
        etudiant=etudiant,
        ecole=ecole_utilisateur,
        annee_scolaire=annee_active
    ).order_by('-date')

    # 🔹 Statistiques
    stats = {
        'total_presents': presences.filter(statut='Présent').count(),
        'total_absents': presences.filter(statut='Absent').count(),
        'total_retards': presences.filter(statut='Retard').count(),
        'total_excuses': presences.filter(statut='Excusé').count(),
    }

    # 🔹 Rendu
    context = {
        'etudiant': etudiant,
        'presences': presences,
        **stats,
        'annee_active': annee_active,
    }

    return render(request, 'dashboard/presences/suivi_presence_eleve.html', context)

# ==========================
# Emplois du temps
# ==========================
# 🔹 LISTE DES EMPLOIS DU TEMPS (FILTRÉ PAR ÉCOLE)
# ==========================
@login_required
def liste_emplois_du_temps(request):
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    annee_active = AnneeScolaire.objects.filter(ecole=ecole, active=True).first()
    toutes_les_classes = Classe.objects.filter(ecole=ecole).order_by('nom_classe')

    classe_id = request.GET.get('classe_id')
    classe_selectionnee = get_object_or_404(Classe, pk=classe_id, ecole=ecole) if classe_id else None
    classes_a_afficher = [classe_selectionnee] if classe_selectionnee else toutes_les_classes

    if not annee_active:
        messages.warning(request, "Veuillez définir une année scolaire active avant de consulter les emplois du temps.")
        return render(request, 'dashboard/emplois_du_temps/emplois_du_temps.html', {
            'annee_active': None,
            'toutes_les_classes': toutes_les_classes,
            'emplois_du_temps_par_classe': {},
            'heures_disponibles': [],
            'jours_semaine': EmploiDuTemps.JourSemaine.values,
        })

    # Colonnes horaires
    plages_horaires = sorted([
        f"{p['heure_debut'].strftime('%H:%M')} - {p['heure_fin'].strftime('%H:%M')}"
        for p in EmploiDuTemps.objects.filter(
            annee_scolaire=annee_active,
            ecole=ecole
        ).values('heure_debut', 'heure_fin').distinct()
    ], key=lambda x: x.split(' - ')[0])

    jours_semaine = EmploiDuTemps.JourSemaine.values
    emplois_par_classe = {}

    for classe in classes_a_afficher:
        edt_items = EmploiDuTemps.objects.filter(
            classe=classe,
            annee_scolaire=annee_active,
            ecole=ecole
        ).select_related('matiere', 'enseignant')

        edt_dict = {jour: {plage: None for plage in plages_horaires} for jour in jours_semaine}
        for item in edt_items:
            plage_str = f"{item.heure_debut.strftime('%H:%M')} - {item.heure_fin.strftime('%H:%M')}"
            if item.jour in edt_dict and plage_str in edt_dict[item.jour]:
                edt_dict[item.jour][plage_str] = item
        emplois_par_classe[classe] = edt_dict

    return render(request, 'dashboard/emplois_du_temps/emplois_du_temps.html', {
        'annee_active': annee_active,
        'toutes_les_classes': toutes_les_classes,
        'classe_selectionnee': classe_selectionnee,
        'emplois_du_temps_par_classe': emplois_par_classe,
        'heures_disponibles': plages_horaires,
        'jours_semaine': jours_semaine,
    })


# ==========================
# 🔹 CRÉATION GÉNÉRALE D'EMPLOI DU TEMPS
# ==========================
@login_required
def creer_emploi_du_temps(request, classe_pk=None):
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    classe = get_object_or_404(Classe, pk=classe_pk, ecole=ecole) if classe_pk else None
    annee_active = AnneeScolaire.objects.filter(ecole=ecole, active=True).first()

    if request.method == 'POST':
        form = EmploiDuTempsForm(request.POST)
        if form.is_valid():
            edt = form.save(commit=False)
            edt.ecole = ecole
            if not edt.annee_scolaire:
                edt.annee_scolaire = annee_active
            edt.save()
            messages.success(request, "Le bloc de cours a été enregistré avec succès !")
            return redirect('liste_emplois_du_temps')
        messages.error(request, "Erreur lors de l'enregistrement. Vérifiez les champs.")
    else:
        initial = {'classe': classe} if classe else {}
        form = EmploiDuTempsForm(initial=initial)

    context = {
        'form': form,
        'classe': classe,
        'action': "Créer" if not request.GET.get('edit') else "Modifier",
    }
    return render(request, 'dashboard/emplois_du_temps/emploi_du_temps_form.html', context)


# ==========================
# 🔹 CRÉATION POUR UNE CLASSE SPÉCIFIQUE
# ==========================
@login_required
def creer_emploi_du_temps_pour_classe(request, classe_id):
    # 🔹 Récupération de l'école de l'utilisateur
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # 🔹 Vérification que la classe appartient à cette école
    classe = get_object_or_404(Classe, pk=classe_id, ecole=ecole)

    # 🔹 Année scolaire active pour cette école
    annee_active = AnneeScolaire.objects.filter(ecole=ecole, active=True).first()
    if not annee_active:
        messages.error(request, "Aucune année scolaire active n’a été trouvée.")
        return redirect('liste_emplois_du_temps')

    if request.method == "POST":
        # 🔹 On passe l'école à notre formulaire pour filtrer les classes, matières et enseignants
        form = EmploiDuTempsForm(request.POST, ecole=ecole)
        if form.is_valid():
            edt = form.save(commit=False)
            edt.ecole = ecole
            edt.classe = classe
            edt.annee_scolaire = annee_active
            edt.save()
            messages.success(request, f"Emploi du temps ajouté pour {classe.nom_classe}.")
            return redirect('liste_emplois_du_temps')
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = EmploiDuTempsForm(initial={'classe': classe}, ecole=ecole)
        form.fields['classe'].widget.attrs['disabled'] = 'disabled'

    return render(request, 'dashboard/emplois_du_temps/form_emploi_du_temps.html', {
        'form': form,
        'action': f"Créer pour {classe.nom_classe}",
        'classe': classe,
    })


# ==========================
# 🔹 MODIFICATION D'UN EMPLOI DU TEMPS
# ==========================
@login_required
def modifier_emploi_du_temps(request, pk):
    # 🔹 Récupération de l'école de l'utilisateur connecté
    ecole_utilisateur = get_user_ecole(request)
    if not ecole_utilisateur:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # 🔹 Récupérer le bloc uniquement s'il appartient à cette école
    bloc_edt = get_object_or_404(EmploiDuTemps, pk=pk, ecole=ecole_utilisateur)

    if request.method == 'POST':
        # 🔹 Passer l'école au formulaire pour filtrer les choix
        form = EmploiDuTempsForm(request.POST, instance=bloc_edt, ecole=ecole_utilisateur)
        if form.is_valid():
            form.save()
            messages.success(request, "Bloc d'emploi du temps mis à jour avec succès !")
            return redirect('liste_emplois_du_temps')
        messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = EmploiDuTempsForm(instance=bloc_edt, ecole=ecole_utilisateur)

    return render(request, 'dashboard/emplois_du_temps/form_emploi_du_temps.html', {
        'form': form,
        'action': 'Modifier'
    })

# ==========================
# 🔹 REDIRECTION MODIFICATION CLASSE
# ==========================
@login_required
def modifier_emploi_du_temps_classe(request, classe_id):
    messages.error(request, "La modification doit se faire bloc par bloc. Utilisez l'icône d'édition dans le tableau.")
    return redirect('liste_emplois_du_temps')
# ==========================
# Saisie de notes
# ==========================
@login_required
def saisir_notes_classe_matiere(request, classe_id, matiere_id):
    """
    Saisie ou modification des notes pour une classe et une matière,
    filtrée par l'école de l'utilisateur connecté.
    """
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    classe = get_object_or_404(Classe, pk=classe_id, ecole=ecole)
    matiere = get_object_or_404(Matiere, pk=matiere_id, ecole=ecole)

    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()
    if not annee_active:
        messages.error(request, "Veuillez définir une année scolaire active pour votre école.")
        return redirect('liste_notes_par_classe_matiere', classe_id=classe_id, matiere_id=matiere_id)

    etudiants = Etudiant.objects.filter(classe=classe, ecole=ecole).order_by('nom', 'prenom')
    notes_existantes = {
        note.etudiant_id: note
        for note in Note.objects.filter(
            etudiant__in=etudiants,
            matiere=matiere,
            annee_scolaire=annee_active,
            ecole=ecole
        )
    }

    if request.method == 'POST':
        periode = request.POST.get('periode_evaluation')
        type_eval = request.POST.get('type_evaluation')
        date_eval = request.POST.get('date_evaluation')

        notes_sauvegardees, erreurs = 0, 0

        for etudiant in etudiants:
            note_str = request.POST.get(f'note_{etudiant.pk}', '').strip()
            if note_str:
                try:
                    note_valeur = float(note_str.replace(',', '.'))
                    Note.objects.update_or_create(
                        etudiant=etudiant,
                        matiere=matiere,
                        periode_evaluation=periode,
                        annee_scolaire=annee_active,
                        ecole=ecole,
                        defaults={
                            'valeur': note_valeur,
                            'type_evaluation': type_eval,
                            'date_evaluation': date_eval,
                        }
                    )
                    notes_sauvegardees += 1
                except ValueError:
                    erreurs += 1
                    messages.warning(request, f"La note de {etudiant.prenom} {etudiant.nom} est invalide.")
                except Exception as e:
                    erreurs += 1
                    messages.error(request, f"Erreur pour {etudiant.prenom} {etudiant.nom} : {e}")

        if notes_sauvegardees:
            messages.success(request, f"{notes_sauvegardees} notes enregistrées ou mises à jour avec succès.")
        if erreurs == 0:
            return redirect('liste_notes_par_classe_matiere', classe_id=classe_id, matiere_id=matiere_id)

    return render(request, 'dashboard/notes/saisir_notes.html', {
        'classe': classe,
        'matiere': matiere,
        'annee_active': annee_active,
        'etudiants': etudiants,
        'notes_existantes': notes_existantes,
    })


# ======================================================
# Modification individuelle d'une note
# ======================================================
@login_required
def modifier_note(request, pk):
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    note = get_object_or_404(Note, pk=pk, ecole=ecole)

    if request.method == 'POST':
        try:
            nouvelle_valeur = request.POST.get('valeur', '').strip()
            note.valeur = float(nouvelle_valeur.replace(',', '.'))
            note.type_evaluation = request.POST.get('type_evaluation')
            note.date_evaluation = request.POST.get('date_evaluation')
            note.save()
            messages.success(request, f"Note de {note.etudiant.prenom} {note.etudiant.nom} mise à jour.")
            return redirect('liste_notes_par_classe_matiere',
                            classe_id=note.etudiant.classe.pk,
                            matiere_id=note.matiere.pk)
        except ValueError:
            messages.error(request, "La valeur de la note doit être un nombre valide.")
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification : {e}")

    return render(request, 'dashboard/notes/modifier_note.html', {'note': note})


# ======================================================
# Suppression individuelle d'une note
# ======================================================
@login_required
def supprimer_note(request, pk):
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    note = get_object_or_404(Note, pk=pk, ecole=ecole)
    classe_id = note.etudiant.classe.pk
    matiere_id = note.matiere.pk

    if request.method == 'POST':
        note.delete()
        messages.success(request, f"Note de {note.etudiant.prenom} {note.etudiant.nom} supprimée.")
        return redirect('liste_notes_par_classe_matiere', classe_id=classe_id, matiere_id=matiere_id)

    return render(request, 'dashboard/notes/confirmer_suppression_note.html', {'note': note})


# ======================================================
# Génération PDF du bulletin scolaire
# ======================================================
@login_required
def generer_bulletin_scolaire(request, etudiant_id, periode):
    etudiant = get_object_or_404(Etudiant, pk=etudiant_id)
    annee_scolaire = AnneeScolaire.objects.filter(active=True).first()
    ecole = EcoleSettings.objects.first()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="bulletin_{etudiant.nom}_{etudiant.prenom}_{periode}.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # --- Récupération des notes ---
    notes_etudiant = Note.objects.filter(
        etudiant=etudiant,
        periode_evaluation=periode,
        annee_scolaire=annee_scolaire
    )

    # Moyenne générale
    moy_gen = sum([n.valeur for n in notes_etudiant]) / len(notes_etudiant) if notes_etudiant else 0.0

    # ------------------- Entête -------------------
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width / 2, height - 40, "BULLETIN DE NOTES")
    if ecole:
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, height - 60, ecole.nom_etablissement)

    p.setFont("Helvetica", 10)
    p.drawString(50, height - 75, f"Année Scolaire : {annee_scolaire.annee if annee_scolaire else 'N/A'}")
    p.drawString(50, height - 90, f"Période : {periode}")

    # ------------------- Infos Élève -------------------
    y = height - 130
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y, "INFORMATIONS ÉLÈVE")
    p.setFont("Helvetica", 10)
    y -= 15
    p.drawString(50, y, f"Nom et prénom : {etudiant.nom} {etudiant.prenom}")
    p.drawString(300, y, f"N° Matricule : {etudiant.numero_matricule or ''}")
    y -= 15
    p.drawString(50, y, f"Classe : {etudiant.classe.nom_classe if etudiant.classe else ''}")
    y -= 30

    # ------------------- Tableau des Notes -------------------
    data = [["Matières", "Note", "Moy Classe", "Mini", "Maxi", "Appréciation"]]
    for note in notes_etudiant:
        data.append([note.matiere.nom, f"{note.valeur:.2f}", "N/A", "N/A", "N/A", ""])

    data.append(["Moyenne Générale", f"{moy_gen:.2f}", "", "", "", ""])

    table = Table(data, colWidths=[120, 60, 60, 50, 50, 150])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.6)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (0, -1), 'Helvetica-Bold'),
    ]))

    y_table = y - len(data)*20
    table.wrapOn(p, width, height)
    table.drawOn(p, 50, y_table)

    # ------------------- Appréciation & Signature -------------------
    y_sign = y_table - 80
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y_sign, "Appréciation du conseil de classe :")
    p.rect(50, y_sign - 50, 400, 50)
    p.drawString(55, y_sign - 30, "Très bon trimestre. Encouragements.")

    p.drawString(460, y_sign, "Cachet et Signature du Chef :")
    p.rect(460, y_sign - 50, 100, 50)

    p.showPage()
    p.save()
    return response

# ======================================================
# Liste et filtrage des notes par classe et matière
# ======================================================
@login_required
def liste_notes_par_classe(request, classe_id=None, matiere_id=None):
    """
    Liste des notes par classe et matière pour l'école de l'utilisateur.
    Si classe_id et matiere_id sont passés en GET ou URL, affiche les notes correspondantes.
    """
    ecole = get_user_ecole(request)
    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # Récupération de l'année scolaire active
    annee_active = AnneeScolaire.objects.filter(ecole=ecole, active=True).first()
    if not annee_active:
        messages.error(request, "Veuillez définir une année scolaire active pour votre école.")
        return render(request, 'dashboard/notes/liste_notes_par_classe.html', {})

    # Priorité : URL parameters > GET parameters
    classe_id = classe_id or request.GET.get('classe_id')
    matiere_id = matiere_id or request.GET.get('matiere_id')

    toutes_les_classes = Classe.objects.filter(ecole=ecole).order_by('nom_classe')
    toutes_les_matieres = Matiere.objects.filter(ecole=ecole).order_by('nom')

    classe = None
    matiere = None
    etudiants_avec_notes = {}

    if classe_id and matiere_id:
        classe = get_object_or_404(Classe, pk=classe_id, ecole=ecole)
        matiere = get_object_or_404(Matiere, pk=matiere_id, ecole=ecole)

        # Étudiants de la classe pour l'année scolaire active
        etudiants = Etudiant.objects.filter(
            classe=classe,
            ecole=ecole,
            annee_scolaire_inscription=annee_active
        ).order_by('nom', 'prenom')

        # Récupérer toutes les notes en une seule requête
        notes = Note.objects.filter(
            etudiant__in=etudiants,
            matiere=matiere,
            annee_scolaire=annee_active,
            ecole=ecole
        ).select_related('etudiant').order_by('etudiant__nom', 'date_evaluation')

        # Construire un dictionnaire {étudiant: [notes]}
        etudiants_avec_notes = {etudiant: [] for etudiant in etudiants}
        for note in notes:
            etudiants_avec_notes[note.etudiant].append(note)

    context = {
        'classe': classe,
        'matiere': matiere,
        'annee_active': annee_active,
        'toutes_les_classes': toutes_les_classes,
        'toutes_les_matieres': toutes_les_matieres,
        'etudiants_avec_notes': etudiants_avec_notes,
    }

    return render(request, 'dashboard/notes/liste_notes_par_classe.html', context)

# ======================================================
# EXPORT DES NOTES VERS EXCEL
# ======================================================
@login_required
def export_notes_excel(request):
    """
    Exporte les notes filtrées par classe, matière et période au format Excel.
    Paramètres GET : classe_id, matiere_id, periode
    """
    classe_id = request.GET.get('classe_id')
    matiere_id = request.GET.get('matiere_id')
    periode = request.GET.get('periode', 'Toutes')

    annee_scolaire = AnneeScolaire.objects.filter(active=True).first()
    if not annee_scolaire:
        messages.error(request, "Aucune année scolaire active trouvée.")
        return redirect('liste_notes_par_classe_matiere')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="notes_{periode}_{annee_scolaire.annee}.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Notes {periode}"

    headers = ["MATRICULE", "NOM", "PRENOM", "CLASSE", "MATIERE", "NOTE", "PERIODE", "ANNEE_SCOLAIRE"]
    sheet.append(headers)

    notes = Note.objects.filter(annee_scolaire=annee_scolaire)
    if classe_id:
        notes = notes.filter(etudiant__classe_id=classe_id)
    if matiere_id:
        notes = notes.filter(matiere_id=matiere_id)
    if periode != 'Toutes':
        notes = notes.filter(periode_evaluation=periode)

    notes = notes.select_related('etudiant', 'matiere', 'etudiant__classe').order_by('etudiant__nom', 'etudiant__prenom')

    for note in notes:
        sheet.append([
            note.etudiant.numero_matricule,
            note.etudiant.nom,
            note.etudiant.prenom,
            note.etudiant.classe.nom_classe if note.etudiant.classe else 'N/A',
            note.matiere.nom,
            note.valeur,
            note.periode_evaluation,
            note.annee_scolaire.annee
        ])

    workbook.save(response)
    return response


# ======================================================
# IMPORT DES NOTES DEPUIS EXCEL
# ======================================================
@login_required
def import_notes_excel(request):
    """
    Importation de notes depuis un fichier Excel.
    Les colonnes doivent correspondre à l'export.
    """
    if request.method == 'POST' and 'file' in request.FILES:
        excel_file = request.FILES['file']

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, "Veuillez télécharger un fichier Excel valide (.xlsx ou .xls).")
            return redirect('liste_notes_par_classe_matiere')

        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            rows = list(sheet.iter_rows(min_row=2, values_only=True))

            notes_creees, notes_mises_a_jour, erreurs = 0, 0, []

            annee_active = AnneeScolaire.objects.filter(active=True).first()
            if not annee_active:
                raise Exception("Aucune année scolaire active n'est définie.")

            for row in rows:
                if not row or not row[0]:
                    continue
                try:
                    matricule, nom, prenom, classe_nom, matiere_nom, valeur, periode, _ = row[:8]
                    etudiant = Etudiant.objects.get(numero_matricule=matricule)
                    matiere = Matiere.objects.get(nom=matiere_nom)
                    note_valeur = float(str(valeur).replace(',', '.'))

                    note_obj, created = Note.objects.update_or_create(
                        etudiant=etudiant,
                        matiere=matiere,
                        periode_evaluation=periode,
                        annee_scolaire=annee_active,
                        defaults={'valeur': note_valeur}
                    )
                    if created:
                        notes_creees += 1
                    else:
                        notes_mises_a_jour += 1
                except Etudiant.DoesNotExist:
                    erreurs.append(f"Matricule introuvable: {matricule}")
                except Matiere.DoesNotExist:
                    erreurs.append(f"Matière introuvable: {matiere_nom}")
                except ValueError:
                    erreurs.append(f"Note invalide pour {matricule} ({valeur})")
                except Exception as e:
                    erreurs.append(f"Erreur inconnue pour {matricule}: {e}")

            if erreurs:
                messages.warning(request, f"Importation terminée : {notes_creees} créations, {notes_mises_a_jour} mises à jour, {len(erreurs)} erreurs.")
                print("Erreurs d'importation:", erreurs)
            else:
                messages.success(request, f"Importation réussie : {notes_creees} nouvelles notes et {notes_mises_a_jour} mises à jour.")
        except Exception as e:
            messages.error(request, f"Erreur lors du traitement du fichier : {e}")

    return render(request, 'dashboard/notes/import_notes.html', {'page_title': "Importer des Notes"})


# ======================================================
# CONFIGURATION DE L'ÉCOLE
@login_required
def config_ecole_view(request):
    """
    Gérer la configuration de l'école pour l'utilisateur connecté.
    Chaque utilisateur ne peut accéder qu'à sa propre école.
    """
    # --- Récupérer l'école de l'utilisateur ---
    ecole = None
    if hasattr(request.user, 'profile') and request.user.profile.ecole:
        ecole = request.user.profile.ecole
    elif hasattr(request.user, 'enseignant') and request.user.enseignant.ecole:
        ecole = request.user.enseignant.ecole
    elif hasattr(request.user, 'etudiant') and request.user.etudiant.ecole:
        ecole = request.user.etudiant.ecole

    if not ecole:
        messages.error(request, "Vous n'êtes associé à aucune école.")
        return redirect('dashboard_accueil')

    # --- Récupérer la configuration existante de l'école ---
    settings = get_object_or_404(EcoleSettings, id=ecole.id)

    # --- Gestion du formulaire POST ---
    if request.method == 'POST':
        form = EcoleSettingsForm(request.POST, request.FILES, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                "Paramètres de l'école mis à jour avec succès."
            )
            return redirect('ecole_settings')
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = EcoleSettingsForm(instance=settings)

    # --- Rendu du template ---
    context = {
        'form': form,
        'settings': settings,
        'page_title': "Paramètres de l'École",
    }
    return render(request, 'dashboard/settings/config_ecole.html', context)

# ======================================================
# CARTE SCOLAIRE (HTML ou PDF)
# ======================================================

@login_required
def carte_scolaire(request, etudiant_id, pdf=False):
    """
    Génère la carte scolaire d'un élève,
    en s'assurant que l'élève appartient à l'école de l'utilisateur connecté.
    """

    # 1️⃣ Identification de l'école de l'utilisateur connecté
    ecole_user = getattr(getattr(request.user, "profile", None), "ecole", None)
    if not ecole_user:
        messages.error(request, "⚠️ Vous n'êtes associé à aucune école.")
        return HttpResponseForbidden("Accès refusé : aucune école associée à ce compte.")

    # 2️⃣ Récupération sécurisée de l'étudiant appartenant à cette école
    etudiant = get_object_or_404(Etudiant, pk=etudiant_id, ecole=ecole_user)

    # 3️⃣ Récupération du contexte scolaire
    ecole = ecole_user
    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole).first()
    if not annee_active:
        messages.warning(request, "Aucune année scolaire active trouvée pour votre école.")
    
    # 4️⃣ Génération du QR Code (Base64)
    qr_url = request.build_absolute_uri(f"/dashboard/verifier_etudiant/{etudiant.pk}/")
    qr = qrcode.QRCode(box_size=6, border=2)
    qr.add_data(qr_url)
    qr.make(fit=True)
    img_qr = qr.make_image(fill_color="black", back_color="white")
    qr_buffer = io.BytesIO()
    img_qr.save(qr_buffer, format="PNG")
    qr_base64 = "data:image/png;base64," + base64.b64encode(qr_buffer.getvalue()).decode()

    # 5️⃣ Préparation du contexte
    context = {
        "etudiant": etudiant,
        "ecole": ecole,
        "annee_scolaire": etudiant.annee_scolaire_inscription or annee_active,
        "qr_code": qr_base64,
    }

    # 6️⃣ Génération PDF (si demandé)
    if pdf:
        return PDFTemplateResponse(
            request=request,
            template='dashboard/cartes/carte_scolaire.html',
            filename=f'carte_{etudiant.numero_matricule}.pdf',
            context=context,
            show_content_in_browser=True,
            cmd_options={'quiet': True}
        )

    # 7️⃣ Sinon, affichage HTML classique
    return render(request, 'dashboard/cartes/carte_scolaire.html', context)


# ======================================================
# VÉRIFICATION D'ÉTUDIANT VIA QR
# ======================================================
@login_required
def verifier_etudiant(request, etudiant_id):
    """
    Vérifie et affiche les informations d’un élève avec QR code,
    limité à l’école de l’utilisateur connecté (multi-école sécurisé).
    """

    # 1️⃣ Récupération de l'école associée à l'utilisateur
    ecole_user = getattr(getattr(request.user, "profile", None), "ecole", None)
    if not ecole_user:
        messages.error(request, "⚠️ Vous n'êtes associé à aucune école.")
        return render(request, "dashboard/erreurs/acces_refuse.html", status=403)

    # 2️⃣ Récupération sécurisée de l'élève appartenant à cette école
    etudiant = get_object_or_404(Etudiant, pk=etudiant_id, ecole=ecole_user)

    # 3️⃣ Récupération de l’année scolaire active de cette école
    annee_active = AnneeScolaire.objects.filter(active=True, ecole=ecole_user).first()
    if not annee_active:
        messages.warning(request, "Aucune année scolaire active trouvée pour votre école.")

    # 4️⃣ Génération du QR Code
    qr_url = request.build_absolute_uri()  # L’URL actuelle
    qr = qrcode.QRCode(box_size=6, border=2)
    qr.add_data(qr_url)
    qr.make(fit=True)
    img_qr = qr.make_image(fill_color="black", back_color="white")
    qr_buffer = io.BytesIO()
    img_qr.save(qr_buffer, format="PNG")
    qr_base64 = "data:image/png;base64," + base64.b64encode(qr_buffer.getvalue()).decode()

    # 5️⃣ Contexte du template
    context = {
        "etudiant": etudiant,
        "ecole": ecole_user,
        "annee_scolaire": etudiant.annee_scolaire_inscription or annee_active,
        "qr_code": qr_base64,
    }

    return render(request, "dashboard/cartes/carte_scolaire.html", context)